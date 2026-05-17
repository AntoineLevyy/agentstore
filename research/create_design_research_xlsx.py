import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

out_dir = "/opt/mrhan/projects/agentstore/research"
os.makedirs(out_dir, exist_ok=True)
out = os.path.join(out_dir, "agentstore-ui-design-research.xlsx")

wb = Workbook()
ws = wb.active
ws.title = "UI Design Trends"

trend_rows = [
    ["Discovery UX", "Faceted search is mandatory", "Filter by category, use case, pricing, integrations, model/provider, trust, deployment type.", "AI agents are risky/ambiguous; users need fit + permission clarity before clicking.", "Sticky desktop filters; mobile filter sheet; chips for integrations, free/paid, verified, open-source, human-in-loop.", "Baymard filtering; NN/g facets; Shopify App Store; Slack App Directory"],
    ["Discovery UX", "Intent search over keyword search", "Users search in jobs-to-be-done language, not product taxonomy.", "AgentAppStore can own 'find me an agent to…' rather than just list names.", "Hero search examples; semantic synonyms; autocomplete; suggested tasks; result match reasons.", "Algolia InstantSearch; Perplexity; Product Hunt"],
    ["Trust", "Privacy/permission labels like app stores", "Show data access, publisher verification, security, retention, and action autonomy.", "Agents may touch Gmail, files, code, CRM, customer data — trust is conversion.", "Badges on cards; detail-page Trust & Data panel above CTA; report/claim listing flows.", "Apple privacy labels; Chrome Web Store; Google Workspace Marketplace"],
    ["Homepage", "Curated homepage, not exhaustive catalog", "Guide users through collections, staff picks, trending, role-based paths.", "Raw catalogs feel like databases; curated paths create taste and SEO surfaces.", "Featured agents, trending this week, best for founders, open-source agents, enterprise-ready agents.", "Apple App Store editorial; Product Hunt; Raycast Store"],
    ["Category Pages", "Category pages as SEO landing pages", "Category page should explain the use case, rank top agents, compare subtypes, answer FAQs.", "Category pages are traffic engines and decision hubs.", "Hero + top picks + filters + subcategory chips + comparison criteria + FAQ + trust module.", "G2, Capterra, Zapier app directory, Shopify App Store categories"],
    ["Comparison", "Native compare UX", "Users need side-by-side evaluation across pricing, integrations, privacy, reviews, and setup time.", "AI agent capabilities overlap; comparison reduces uncertainty.", "Compare checkbox on cards; 2–4 item drawer; indexable /compare/a-vs-b pages.", "G2 compare; Capterra compare; AlternativeTo"],
    ["Cards", "Cards must communicate job-to-be-done", "Each card says what the agent actually does, not vague AI claims.", "Clear outcomes improve scan speed and CTR.", "Name, one-sentence task outcome, category, price, integrations, trust badge, rating, CTA.", "Intercom JTBD; Zapier app listings; Material cards"],
    ["Typography", "High-legibility UI fonts + tabular data", "Modern sans with strong hierarchy, generous line-height, tabular numerals.", "Directories are comparison-heavy; bad type causes fatigue.", "Inter/Geist/IBM Plex/Manrope/Plus Jakarta; 48–64 H1; 16–18 body; 12–14 metadata.", "Vercel Geist; Inter; Material type scale"],
    ["Visual Style", "Functional AI-native depth", "Soft gradients, subtle glass, layered cards, but keep result grids solid/readable.", "Modern without looking like generic AI slop.", "Gradient hero/search; solid listing cards; subtle hover lift; selected compare border.", "Linear; Vercel; Framer; Mobbin patterns"],
    ["Shapes", "Rounded card/chip system", "Use consistent radii for app icons, cards, chips, buttons, panels.", "Dense marketplaces feel calmer and more tappable.", "Cards 16–24px; buttons 10–14px; chips full-pill; app icons 20–28% corner radius.", "Apple App Store; Material shape; Shopify Polaris"],
    ["Accessibility", "A11y is part of the design system", "Keyboard filters/search, visible focus, semantic links/buttons, contrast, result count announcements.", "Directory UX relies on controls that often break accessibility.", "ARIA result counts; keyboard-navigable suggestions; real controls; WCAG AA color tokens.", "WCAG 2.2; WebAIM; GOV.UK Design System"],
    ["Performance", "Performance is UX", "Fast search, stable grids, optimized icons, static category pages, lazy loading.", "Slow catalog search destroys trust and SEO.", "Pre-render categories; virtualize large grids later; optimize remote icons; skeleton states.", "Core Web Vitals; Next.js optimization docs"],
    ["Mobile", "Mobile filters need bottom sheets", "Desktop sidebars become sticky search, horizontal chips, filter sheet, single-column cards.", "Mobile discovery often precedes desktop install/config.", "Sticky search; filter bottom sheet; active-filter chips; sort visible; card CTAs thumb-safe.", "Material bottom sheets; iOS HIG; Baymard mobile ecommerce"],
    ["Messaging", "Move from AI magic to practical outcomes", "Replace hype with specific saved time, automation, accuracy, workflow fit.", "Users are skeptical of AI directories; specificity builds trust.", "'Drafts customer replies from help docs' not 'AI-powered support assistant'.", "NN/g AI UX; Zapier AI; Intercom Fin"],
]
ws.append(["Area", "Trend or Rule", "What it means", "Why it matters for AgentAppStore", "Concrete site application", "References / examples"])
for r in trend_rows:
    ws.append(r)

ws2 = wb.create_sheet("Category Trends")
ws2.append(["AgentAppStore category", "User moment", "Visual style direction", "Typography / shape cues", "Content modules to add", "Trust / proof needed", "SEO content angle", "Examples / references"])
cat_rows = [
    ["Productivity", "Overwhelmed user wants fewer context switches and clearer next actions.", "Calm task dashboards, checklists, progress cards, subtle gradients.", "Rounded cards, soft shadows, status pills, task-list motifs.", "Daily planning agents; meeting/email/calendar comparisons; workflow examples; time-saved calculator.", "Gmail/Calendar/Slack/Notion permissions; privacy for email/calendar access.", "Best AI productivity agents; AI calendar assistants; AI meeting note takers.", "Motion, Reclaim, Notion AI, Todoist AI, Granola, Otter"],
    ["Developer Tools", "Developer needs to ship, debug, review, or understand code faster.", "Dark IDE/terminal panels, diffs, repo graphs, code previews.", "Monospace accents, syntax colors, diff markers, sharper cards.", "Use cases: coding/debugging/review/tests/migration; IDE filters; language support; issue-to-PR examples.", "GitHub permissions, local vs cloud code processing, SOC2/security, secrets handling.", "Best AI coding agents; AI code review tools; AI debugging agents.", "Cursor, GitHub Copilot, Devin, Windsurf, Sourcegraph Cody, Aider"],
    ["Writing", "Marketer/founder needs quality copy, SEO, scripts, docs, or brand-safe content.", "Editorial writing studio, content calendar, before/after copy blocks.", "Expressive headings, annotation highlights, document cards.", "By channel: blogs/email/social/docs; brand voice modules; plagiarism checks; example transformations.", "Brand safety, originality, source/citation controls, publication integrations.", "Best AI writing agents; AI SEO writers; AI copywriting agents.", "Jasper, Copy.ai, Writer, Typeface, Surfer SEO"],
    ["Research", "User needs reliable synthesis with citations and freshness.", "Source cards, citation trails, knowledge graphs, document stacks.", "Serif credibility accents, citation badges, highlighted excerpts.", "Source-type filters; citation-quality comparison; deep-research templates; export options.", "Source transparency, hallucination mitigation, date freshness, inspectable sources.", "Best AI research agents; AI agents with citations; AI literature review tools.", "Perplexity, Elicit, Consensus, SciSpace, NotebookLM"],
    ["Data Analysis", "User wants answers from data without hand-writing every query/dashboard.", "BI dashboards, conversational analytics, query result cards.", "Tabular numerals, SQL/code accents, compact chart tiles.", "Data-source filters; NL-to-SQL examples; dashboard generation demos; governance explainer.", "Permission-aware querying, lineage, accuracy validation, RBAC, tenant isolation.", "Best AI data analytics agents; natural language analytics tools; AI SQL agents.", "ThoughtSpot, Tableau Pulse, Power BI Copilot, Hex, Julius, Seek AI"],
    ["Customer Support", "Team needs to reduce tickets and answer customers accurately with handoff.", "Support inbox, chat-first UI, help-center article cards, ticket queues.", "Rounded chat bubbles, priority badges, CSAT cards, friendly microcopy.", "Channel filters; deflection benchmarks; human handoff examples; KB sync guide.", "Approved-doc grounding, escalation rules, audit logs, multilingual, PII handling.", "Best AI customer support agents; AI helpdesk automation; AI support chatbot.", "Intercom Fin, Zendesk AI, Ada, Gorgias, Forethought, Decagon, Sierra"],
    ["Creative", "Creator needs to generate/edit/remix/publish visual/audio/video/story artifacts.", "Gallery-first, output previews, creator portfolios, timelines.", "Large display type, media tiles, waveforms/timelines, irregular creative shapes.", "By medium filters; prompt-to-output galleries; style comparisons; commercial rights section.", "Copyright/IP rights, model training transparency, watermarking, commercial-use terms.", "Best AI creative agents; AI video agents; AI image generation agents; AI music agents.", "Midjourney, Runway, Adobe Firefly, Canva AI, Suno, ElevenLabs, Pika"],
    ["Education", "Student/teacher wants tutoring, lesson plans, grading, or study support.", "Warm notebook/progress-path interface, curriculum maps.", "Friendly rounded type, hand annotations, progress rings, badges.", "By learner filters; subject filters; lesson-plan examples; academic integrity guide.", "Age safety, FERPA/COPPA/GDPR notes, anti-cheating controls, expert-reviewed content.", "Best AI education agents; AI tutoring agents; AI lesson planning tools.", "Khanmigo, Duolingo Max, Quizlet AI, MagicSchool, Socratic"],
    ["Finance", "Business user needs forecasting, analysis, bookkeeping, FP&A, tax workflows.", "High-trust fintech, dark navy/green, charts, forecast panels.", "Tabular numerals, compact data cards, line charts, secure badges.", "Budgeting/bookkeeping/FP&A/tax use cases; risk disclosures; spreadsheet export; integrations.", "Regulatory disclaimers, bank integration safety, data security, audit trails.", "Best AI finance agents; AI bookkeeping agents; AI financial analysis tools.", "Intuit Assist, Ramp Intelligence, Brex AI, FinChat, Datarails"],
    ["Sales & Marketing", "Team wants pipeline, outreach, campaign content, SEO, and less CRM admin.", "CRM pipeline + campaign board hybrid; lead cards; ad creative grids.", "Sharp cards, conversion arrows, numeric emphasis, colorful channel tags.", "Outbound/inbound/account research/email/ads/SEO filters; CRM matrix; before/after examples.", "Deliverability, compliance, CRM write permissions, attribution, brand safety.", "Best AI sales agents; AI SDR agents; AI marketing agents; AI campaign automation.", "Clay, Apollo AI, Salesforce Einstein, HubSpot AI, Lavender, Regie.ai, Jasper"],
    ["Health & Wellness", "Consumer wants coaching, habit change, workout/nutrition/sleep support.", "Clean wellness, soft greens/blues, habit streaks, coach cards.", "Rounded type, rings, streak bars, biomarker cards, gentle icons.", "By goal filters; device/wearable compatibility; sample plans; safety disclaimers.", "Medical disclaimers, expert credentials, HIPAA where relevant, diagnosis limits.", "Best AI fitness agents; AI health coach; AI nutrition planner.", "Whoop Coach, Fitbit AI, Apple Health, MyFitnessPal AI, Noom"],
    ["Shopping & Home", "Consumer wants best product/deal/home choice without endless browsing.", "Retail grids, recommendation quizzes, comparison panels, deal badges.", "Price tags, star ratings, swatches, checkmarks, product cards.", "Fashion/electronics/gifts/groceries/home filters; compare matrix; quiz; price history.", "Affiliate disclosure, review authenticity, price freshness, return policies.", "Best AI shopping assistants; AI product recommendation agents; AI deal finder.", "Amazon Rufus, Klarna AI, Google Shopping AI, Perplexity Shopping"],
    ["Travel & Lifestyle", "Traveler wants itinerary around budget, preferences, timing, constraints.", "Destination imagery, itinerary maps, booking cards, day timelines.", "Editorial travel headings, map pins, route lines, itinerary cards.", "Family/business/solo/budget/luxury filters; itinerary demos; price tracking; local recs.", "Live pricing/availability, cancellation clarity, source links, affiliate transparency.", "Best AI travel agents; AI trip planner; AI itinerary generator.", "Expedia Romie, Booking.com AI Trip Planner, Kayak AI, Mindtrip, Layla"],
    ["Relationships & Social", "User wants help with dating, gifts, communication, events, or companionship.", "Warm intimate chat/journal/plan interface, social cards.", "Avatar circles, mood colors, diary cards, warm rounded fonts.", "By relationship type; gift/event planners; conversation coaching; memory controls.", "Privacy/memory controls, emotional safety, moderation, clear AI disclosure.", "Best AI companion apps; AI dating coach; AI gift planner; AI social assistant.", "Pi, Character.AI, Replika, Nomi, Rosebud, Kindroid"],
    ["Learning & Skills", "Consumer wants to learn cooking, music, language, hobbies, or practical skills.", "Studio/classroom hybrid; practice cards; progress streaks.", "Friendly rounded type, lesson cards, achievement badges, hand-drawn cues.", "By skill filters; practice plans; coach feedback examples; progress paths.", "Expert-reviewed content, safety for physical skills, age appropriateness.", "AI language tutor; AI cooking coach; AI music tutor; best AI skill-learning agents.", "Duolingo Max, Yousician, Khanmigo, Quizlet, Skillshare/learning assistants"],
    ["Parenting & Family", "Parent wants help with baby tracking, child safety, schedules, family logistics.", "Calm family OS, routine cards, shared calendars, safety checklists.", "Soft rounded shapes, pastel accents, checklist and timeline motifs.", "Age-stage filters; routine generators; child safety guide; family calendar workflows.", "Child data privacy, pediatric disclaimers, emergency boundaries, COPPA/GDPR-K notes.", "Best AI parenting apps; AI family assistant; baby tracking AI.", "Baby trackers, family calendar apps, school communication assistants"],
    ["Legal & Admin", "User needs contracts, job search, immigration, tax, or admin paperwork help.", "Document-first, redline comparison, secure vault, admin checklist.", "Serif/legal accents, clause cards, stamps, redline highlights.", "By task filters; jurisdiction filters; clause extraction; attorney/pro review disclaimers.", "Not legal advice, confidentiality, jurisdiction boundaries, citation accuracy, security.", "Best AI legal agents; AI contract review; AI immigration assistant; AI tax assistant.", "Harvey, Spellbook, CoCounsel, Lexis+ AI, Luminance, Ironclad AI"],
    ["Personal Finance", "Consumer wants budgeting, saving, investing, credit, and money clarity.", "Consumer fintech, secure but friendly, budget cards and forecast charts.", "Tabular numerals, progress rings, goal cards, green/blue trust palette.", "Budgeting/saving/investing/credit filters; calculators; bank-linking explainer; risk disclosures.", "Bank data safety, no-advice limits, regulatory disclaimers, pricing transparency.", "Best AI budgeting apps; AI personal finance assistant; AI investment assistant.", "Cleo, Rocket Money, Intuit Assist, Monarch, YNAB-like assistants"],
]
for r in cat_rows:
    ws2.append(r)

ws3 = wb.create_sheet("Design System Direction")
ws3.append(["Decision area", "Recommendation", "Concrete tokens / implementation", "Reason"])
for r in [
    ["Positioning", "AgentAppStore should be an intent-based discovery engine, not a static agent list.", "Hero search as primary artifact; category pages as decision hubs; content around jobs-to-be-done.", "This creates stronger user value and SEO moat."],
    ["Font", "Use Geist/Inter stack now; consider display treatment through weight/spacing not exotic font.", "--font-sans Geist/Inter; H1 64–80 desktop, 38–44 mobile; body 15–17; metadata 11–13.", "Modern, legible, compatible with current Next setup."],
    ["Color", "Keep dark premium base but add category accent system.", "Base #08090a/#0f1011; brand #5e6ad2; category accents for work/consumer/safety/creative/dev.", "Current dark Linear-ish system is good but too uniform."],
    ["Shape", "Increase marketplace warmth with larger radii on cards, panels, and chips.", "Cards 16–24px; controls 12–16px; chips pill; icons 22% radius.", "Makes dense catalog feel consumer-quality."],
    ["Cards", "Agent cards need proof and fit signals.", "Add integrations/tools, pricing, rating, verified/trust badge, match reason, 'best for'.", "Agent quality cannot be evaluated from name+tagline alone."],
    ["Category pages", "Rebuild into content-rich landing pages.", "Top-picks strip, category intro, sub-use-case chips, trust module, FAQ, comparison criteria.", "Category pages can fuel SEO and richer browsing."],
    ["Trust", "Trust modules must be visible near CTA.", "Permissions, data access, publisher, freshness, limitation, report/claim listing.", "AI-agent directories need app-store-level safety cues."],
    ["Mobile", "Design mobile independently.", "Sticky search, horizontal chips, filter bottom sheet, single-column high-signal cards.", "Mobile discovery differs from desktop evaluation."],
]:
    ws3.append(r)

ws4 = wb.create_sheet("Sources")
ws4.append(["Source", "URL", "Use"])
for r in [
    ["Baymard filtering research", "https://baymard.com/research/product-lists-filtering", "Facets, sorting, product-list usability"],
    ["Nielsen Norman Group - filters vs facets", "https://www.nngroup.com/articles/filters-vs-facets/", "Filter/facet guidance"],
    ["Apple App Store privacy labels", "https://developer.apple.com/app-store/app-privacy-details/", "Trust/privacy labels"],
    ["WCAG 2.2", "https://www.w3.org/TR/WCAG22/", "Accessibility baseline"],
    ["Material Design cards", "https://m3.material.io/components/cards/overview", "Card structure and interaction"],
    ["Material shape", "https://m3.material.io/styles/shape/overview", "Radius/shape system"],
    ["Vercel Geist", "https://vercel.com/font", "Typography"],
    ["Inter", "https://rsms.me/inter/", "Typography"],
    ["Google structured data", "https://developers.google.com/search/docs/appearance/structured-data/intro-structured-data", "SEO schema"],
    ["Schema.org SoftwareApplication", "https://schema.org/SoftwareApplication", "Listing schema"],
    ["Shopify App Store", "https://apps.shopify.com/", "Marketplace/category UX"],
    ["Slack App Directory", "https://slack.com/apps", "App directory UX"],
    ["Raycast Store", "https://www.raycast.com/store", "Developer tool marketplace UX"],
    ["G2 categories", "https://www.g2.com/categories", "Category/comparison SEO pages"],
    ["Capterra", "https://www.capterra.com/", "Software directory comparison UX"],
]:
    ws4.append(r)

header_fill = PatternFill("solid", fgColor="111827")
header_font = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="E5E7EB")
for wsx in wb.worksheets:
    wsx.freeze_panes = "A2"
    for cell in wsx[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = Border(bottom=thin)
    for row in wsx.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=Side(style="hair", color="E5E7EB"))
    widths = {
        "UI Design Trends": [18, 28, 42, 44, 48, 36],
        "Category Trends": [22, 42, 34, 32, 48, 42, 36, 32],
        "Design System Direction": [22, 48, 48, 38],
        "Sources": [34, 56, 42],
    }[wsx.title]
    for i, w in enumerate(widths, start=1):
        wsx.column_dimensions[get_column_letter(i)].width = w
    end_col = get_column_letter(wsx.max_column)
    ref = f"A1:{end_col}{wsx.max_row}"
    table_name = wsx.title.replace(" ", "")[:20]
    tab = Table(displayName=table_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    wsx.add_table(tab)

wb.save(out)
print(out)
